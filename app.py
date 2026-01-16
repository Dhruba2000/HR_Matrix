import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO


def generate_output(df):
    # Clean column names
    df.columns = [c.strip().lower() for c in df.columns]

    matrix_names = {(0,0):[], (0,1):[], (1,0):[], (1,1):[]}

    for _, row in df.iterrows():
        att, disc, name = int(row['attendance']), int(row['discipline']), str(row['name'])
        if att == 0 and disc == 1: matrix_names[(0,0)].append(name)
        elif att == 1 and disc == 1: matrix_names[(0,1)].append(name)
        elif att == 0 and disc == 0: matrix_names[(1,0)].append(name)
        elif att == 1 and disc == 0: matrix_names[(1,1)].append(name)

    def classify(r):
        att, disc = r['attendance'], r['discipline']
        if att == 0 and disc == 1: return 'Low Attendance, Good Discipline'
        elif att == 1 and disc == 1: return 'Good Attendance, Good Discipline'
        elif att == 0 and disc == 0: return 'Low Attendance, Poor Discipline'
        else: return 'Good Attendance, Poor Discipline'

    df['classification'] = df.apply(classify, axis=1)

    output = BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Employee Details")
    for r in dataframe_to_rows(df, index=False, header=True): ws1.append(r)

    # Matrix Sheet
    ws2 = wb.create_sheet("Classification Matrix")
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws2.merge_cells('B1:C1'); ws2['B1']="Attendance"; ws2['B1'].alignment=Alignment(horizontal='center'); ws2['B1'].font=Font(bold=True)
    ws2['B2'], ws2['C2'] = '0', '1'
    for cell in ['B2','C2']: ws2[cell].alignment=Alignment(horizontal='center'); ws2[cell].font=Font(bold=True)
    ws2.merge_cells('A1:A2'); ws2['A1']="Discipline"; ws2['A1'].alignment=Alignment(horizontal='center'); ws2['A1'].font=Font(bold=True)

    ws2['A3'], ws2['A4'] = '1','0'
    ws2['B3'], ws2['C3'] = '\n'.join(matrix_names[(0,0)]), '\n'.join(matrix_names[(0,1)])
    ws2['B4'], ws2['C4'] = '\n'.join(matrix_names[(1,0)]), '\n'.join(matrix_names[(1,1)])
    for cell in ['B3','C3','B4','C4']: ws2[cell].alignment=Alignment(wrap_text=True, horizontal='center')

    # Summary Sheet
    summary = pd.DataFrame({
        'Category': ['Total','Good Attendance','Low Attendance','Good Discipline','Poor Discipline','Both Good (1,1)','Both Poor (0,0)'],
        'Count': [
            len(df),
            len(df[df['attendance']==1]),
            len(df[df['attendance']==0]),
            len(df[df['discipline']==1]),
            len(df[df['discipline']==0]),
            len(df[(df['attendance']==1)&(df['discipline']==1)]),
            len(df[(df['attendance']==0)&(df['discipline']==0)])
        ]
    })
    ws3 = wb.create_sheet("Summary")
    for r in dataframe_to_rows(summary, index=False, header=True): ws3.append(r)

    wb.save(output)
    output.seek(0)
    return output


# ===== STREAMLIT UI =====

st.title("📊 Employee Classification Matrix")
st.write("Upload Excel file containing columns: **Name, Discipline, Attendance**")

uploaded = st.file_uploader("Upload Excel File", type=['xlsx'])

if uploaded:
    df = pd.read_excel(uploaded)
    st.success("File uploaded successfully!")
    st.dataframe(df)

    if st.button("Generate Matrix"):
        result_file = generate_output(df)
        st.success("Matrix generated!")

        st.download_button(
            label="⬇ Download Output Excel",
            data=result_file,
            file_name="Employee_classification_output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
